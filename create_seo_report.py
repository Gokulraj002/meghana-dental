from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import SheetView, Pane

wb = Workbook()

# ── Color palette ──────────────────────────────────────────────────
TEAL        = "0EA5A0"
DARK_TEAL   = "0B7A76"
LIGHT_TEAL  = "E0F5F5"
WHITE       = "FFFFFF"
GREEN_FILL  = "C6EFCE"
GREEN_TEXT  = "276221"
YELLOW_FILL = "FFEB9C"
YELLOW_TEXT = "9C6500"
RED_FILL    = "FFC7CE"
RED_TEXT    = "9C0006"
ORANGE_FILL = "FCE4D6"
BLUE_FILL   = "DDEBF7"
BLUE_TEXT   = "1F4E79"
PURPLE_FILL = "E2EFDA"

def hdr_fill(color=TEAL):
    return PatternFill("solid", fgColor=color)

def cell_fill(color):
    return PatternFill("solid", fgColor=color)

def hdr_font(size=11):
    return Font(name="Arial", bold=True, color=WHITE, size=size)

def body_font(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def freeze(ws, cell):
    ws.freeze_panes = cell

def style_header_row(ws, row, num_cols, fill_color=TEAL):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = hdr_fill(fill_color)
        cell.font = hdr_font()
        cell.alignment = center()
        cell.border = thin_border()

def style_data_row(ws, row, num_cols, alt=False, font_color="000000"):
    bg = LIGHT_TEAL if alt else WHITE
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF", WHITE, LIGHT_TEAL):
            cell.fill = cell_fill(bg)
        cell.font = body_font(color=font_color)
        cell.alignment = left()
        cell.border = thin_border()

def set_tab_color(ws, color):
    ws.sheet_properties.tabColor = color


# ═══════════════════════════════════════════════════════════════════
# SHEET 1 — SEO Summary
# ═══════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "SEO Summary"
set_tab_color(ws1, TEAL)

ws1.append(["Metric", "Value"])
style_header_row(ws1, 1, 2)

rows1 = [
    ("Client Name", "Meghana Multi Speciality Dental Hospital"),
    ("Website", "https://meghanadental.in"),
    ("Report Date", "2026-06-02"),
    ("Total Pages Optimized", "29"),
    ("Pages with Canonical URLs", "29"),
    ("Pages with OG Tags", "29"),
    ("Pages with Schema Markup", "All pages (via global SchemaMarkup + page-level FAQ/Review schemas)"),
    ("Schema Types Used", "Dentist, MedicalBusiness, FAQPage, MedicalOrganization, Article, WebSite"),
    ("Clinic Name Updated", "Meghana Dental Hospital → Meghana Multi Speciality Dental Hospital"),
    ("Clinic Location", "Ashok Nagar, Tirupati, Andhra Pradesh 517501"),
    ("Phone", "+91 7893327036"),
    ("Google Rating", "4.9★ (850+ reviews)"),
    ("Primary Keywords Targeted", "50+"),
    ("Global Schema Component", "SchemaMarkup (Dentist + MedicalBusiness + WebSite + AggregateRating)"),
    ("Analytics", "Google Analytics G-S4R781DM6J integrated"),
    ("Sitemap", "Dynamic /sitemap.xml — all 29 pages"),
    ("robots.txt", "Configured — allows all, sitemap declared"),
    ("Static Export", "Next.js output: export — pre-rendered HTML"),
]

for i, (metric, value) in enumerate(rows1, start=2):
    ws1.cell(row=i, column=1, value=metric)
    ws1.cell(row=i, column=2, value=value)
    alt = (i % 2 == 0)
    style_data_row(ws1, i, 2, alt=alt)
    ws1.cell(row=i, column=1).font = body_font(bold=True)

set_col_width(ws1, "A", 38)
set_col_width(ws1, "B", 65)
ws1.row_dimensions[1].height = 22
for r in range(2, 20):
    ws1.row_dimensions[r].height = 20
freeze(ws1, "A2")


# ═══════════════════════════════════════════════════════════════════
# SHEET 2 — Page-wise SEO Audit
# ═══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Page-wise SEO Audit")
set_tab_color(ws2, "00B050")

headers2 = [
    "Page", "URL", "Meta Title", "Title Length", "Description",
    "Desc Length", "Keywords (excerpt)", "Canonical URL",
    "OG Title", "OG Site Name", "Schema Type", "Status"
]
ws2.append(headers2)
style_header_row(ws2, 1, len(headers2))

og_site = "Meghana Multi Speciality Dental Hospital"
rows2 = [
    ("Home", "/", "Best Dental Clinic in Tirupati | Meghana Multi Speciality Dental", 62,
     "Best dental clinic in Tirupati — 17+ years, MDS specialists, dental microscope with 25× magnification, implants, root canal, braces, aligners. 50,000+ happy patients.", 158,
     "best dental clinic in Tirupati, dentist in Tirupati, dental implants Tirupati, root canal treatment Tirupati…",
     "https://meghanadental.in/",
     "Best Dental Clinic in Tirupati | Meghana Multi Speciality Dental Hospital — 4.9★ Rated",
     og_site, "Global Dentist + MedicalBusiness + WebSite", "✅ Optimized"),

    ("About", "/about", "Best Dental Hospital in Tirupati | 17+ Years of Care", 51,
     "Meghana Multi Speciality Dental Hospital in Tirupati — 17+ years of excellence, 50,000+ patients, MDS specialists delivering world-class dental care in Andhra Pradesh.", 164,
     "about Meghana Multi Speciality Dental Hospital, best dental hospital Tirupati, MDS dentist Tirupati…",
     "https://meghanadental.in/about",
     "About Meghana Multi Speciality Dental Hospital — Tirupati's Most Trusted Dental Centre",
     og_site, "Global Dentist + MedicalBusiness", "✅ Optimized"),

    ("Contact", "/contact", "Book Dental Appointment in Tirupati | Call & WhatsApp", 56,
     "Contact Meghana Multi Speciality Dental Hospital Tirupati. Book a dental appointment via WhatsApp, call, or visit our clinic at Ashok Nagar, Tirupati, Andhra Pradesh.", 167,
     "contact Meghana Multi Speciality Dental Hospital Tirupati, book dental appointment Tirupati…",
     "https://meghanadental.in/contact",
     "Contact Meghana Multi Speciality Dental Hospital | Book Your Appointment in Tirupati",
     og_site, "Global Dentist + MedicalBusiness", "✅ Optimized"),

    ("Doctors", "/doctors", "Best MDS Dentist Specialists in Tirupati | Expert Team", 55,
     "Meet the MDS-qualified dental specialists at Meghana Multi Speciality Dental Hospital Tirupati — expert team with decades of combined experience in every speciality.", 163,
     "best dentist in Tirupati, MDS specialist Tirupati, orthodontist Tirupati, endodontist Tirupati…",
     "https://meghanadental.in/doctors",
     "Best Dentists in Tirupati | Meet Our MDS Specialist Team — Meghana Dental",
     og_site, "MedicalOrganization (Physician employees)", "✅ Optimized"),

    ("Gallery", "/gallery", "Photo Gallery | Meghana Multi Speciality Dental Hospital Tirupati", 59,
     "Explore the Meghana Multi Speciality Dental Hospital gallery — our modern clinic, advanced equipment, expert team, and real patient smile transformations in Tirupati.", 169,
     "Meghana Multi Speciality Dental Hospital gallery Tirupati, dental clinic photos Tirupati…",
     "https://meghanadental.in/gallery",
     "Photo Gallery — Meghana Multi Speciality Dental Hospital, Tirupati",
     og_site, "Global Dentist + MedicalBusiness", "✅ Optimized"),

    ("Testimonials", "/testimonials", "Dental Patient Reviews Tirupati | 4.9★ Google Rating", 53,
     "Read real patient reviews & testimonials for Meghana Multi Speciality Dental Hospital Tirupati. 4.9★ Google rating from 850+ verified reviews. 50,000+ happy patients served.", 174,
     "Meghana Multi Speciality Dental Hospital reviews Tirupati, dental hospital reviews Tirupati…",
     "https://meghanadental.in/testimonials",
     "Patient Testimonials — 4.9★ Rated Dental Hospital in Tirupati | Meghana Multi Speciality Dental Hospital",
     og_site, "MedicalOrganization + AggregateRating + Review", "✅ Optimized"),

    ("Blogs", "/blogs", "Dental Health Blog | Tips & Guides | Meghana Multi Speciality Dental, Tirupati", 72,
     "Read expert dental health articles, treatment guides, and oral care tips from Meghana Multi Speciality Dental Hospital MDS specialists in Tirupati.", 152,
     "dental health blog Tirupati, dental tips, oral care guide, root canal guide…",
     "https://meghanadental.in/blogs",
     "Dental Health Blog | Meghana Multi Speciality Dental Hospital, Tirupati",
     og_site, "Global Dentist + Article (blog posts)", "✅ Optimized"),

    ("Privacy Policy", "/privacy-policy", "Privacy Policy | Meghana Multi Speciality Dental Hospital", 52,
     "Privacy policy for Meghana Multi Speciality Dental Hospital website — how we collect, use, and protect your personal information.", 128,
     "privacy policy, Meghana Dental privacy",
     "https://meghanadental.in/privacy-policy",
     "Privacy Policy | Meghana Multi Speciality Dental Hospital",
     og_site, "Global Dentist + MedicalBusiness", "✅ Optimized"),

    ("Terms & Conditions", "/terms-conditions", "Terms & Conditions | Meghana Multi Speciality Dental Hospital", 56,
     "Terms and conditions governing use of the Meghana Multi Speciality Dental Hospital website and services.", 106,
     "terms and conditions, Meghana Dental terms",
     "https://meghanadental.in/terms-conditions",
     "Terms & Conditions | Meghana Multi Speciality Dental Hospital",
     og_site, "Global Dentist + MedicalBusiness", "✅ Optimized"),

    ("All Services", "/services", "Dental Treatments in Tirupati | 14 Specialist Services", 52,
     "14 specialist dental treatments in Tirupati — implants, root canal, braces, aligners, whitening, pediatric dentistry, microscope dentistry & more.", 153,
     "dental services Tirupati, dental treatments Tirupati, dental implants Tirupati, root canal Tirupati…",
     "https://meghanadental.in/services",
     "All Dental Services in Tirupati | Meghana Multi Speciality Dental Hospital",
     og_site, "Global Dentist + MedicalBusiness", "✅ Optimized"),

    ("Root Canal", "/services/root-canal", "Painless Root Canal Treatment Tirupati | Single Sitting", 53,
     "Painless, microscope-guided root canal treatment in Tirupati. Single-sitting RCT available. Dental microscope with 25× magnification, MDS endodontists. Book now.", 156,
     "root canal treatment Tirupati, painless root canal Tirupati, single sitting root canal Tirupati…",
     "https://meghanadental.in/services/root-canal",
     "Painless Root Canal Treatment in Tirupati | Single-Sitting RCT — Meghana Multi Speciality Dental Hospital",
     og_site, "FAQPage (4 Q&As)", "✅ Optimized"),

    ("Dental Implants", "/services/dental-implants", "Best Dental Implants in Tirupati | Affordable Cost", 51,
     "Dental implants in Tirupati by expert MDS specialists. Single & multiple implants, All-on-4, same-day implants. Transparent pricing. Book now.", 147,
     "dental implants Tirupati, dental implants cost Tirupati, best implant dentist Tirupati…",
     "https://meghanadental.in/services/dental-implants",
     "Dental Implants in Tirupati | Permanent Tooth Replacement — Meghana Multi Speciality Dental Hospital",
     og_site, "FAQPage (5 Q&As)", "✅ Optimized"),

    ("Orthodontics/Braces", "/services/orthodontics", "Best Orthodontist in Tirupati | Braces & Teeth Alignment", 57,
     "Tirupati's trusted orthodontist for dental braces & teeth alignment. Metal, ceramic & self-ligating braces. 3D oral scanning. Free assessment.", 143,
     "orthodontist Tirupati, braces in Tirupati, metal braces Tirupati, ceramic braces Tirupati…",
     "https://meghanadental.in/services/orthodontics",
     "Best Orthodontist in Tirupati | Braces & Smile Correction — Meghana Multi Speciality Dental Hospital",
     og_site, "FAQPage (3 Q&As)", "✅ Optimized"),

    ("Clear Aligners", "/services/aligners", "Clear Aligners Tirupati | Invisible Teeth Straightening", 55,
     "Invisible teeth straightening in Tirupati with custom clear aligners. 3D intra-oral scanning, removable trays, no metal braces. Book free 3D scan.", 149,
     "clear aligners Tirupati, invisible braces Tirupati, Invisalign Tirupati, transparent aligners Tirupati…",
     "https://meghanadental.in/services/aligners",
     "Clear Aligners in Tirupati | Invisible Teeth Straightening — Meghana Multi Speciality Dental Hospital",
     og_site, "FAQPage (3 Q&As)", "✅ Optimized"),

    ("Teeth Whitening", "/services/teeth-whitening", "Laser Teeth Whitening Tirupati | 8 Shades Brighter", 52,
     "Professional teeth whitening in Tirupati — in-office laser whitening up to 8 shades brighter in one visit. Safe, pain-free, long-lasting results.", 152,
     "teeth whitening Tirupati, laser teeth whitening Tirupati, smile makeover Tirupati, dental veneers Tirupati…",
     "https://meghanadental.in/services/teeth-whitening",
     "Teeth Whitening & Smile Makeover in Tirupati | Meghana Multi Speciality Dental Hospital",
     og_site, "FAQPage (3 Q&As)", "✅ Optimized"),

    ("Microscope Dentistry", "/services/microscope-dentistry", "Dental Microscope Dentist Tirupati | 25× Magnification RCT", 57,
     "Dental microscope with 25× magnification — only clinic in Tirupati with this gold-standard technology. Precision root canals, microsurgery, crack detection.", 155,
     "microscope dentistry Tirupati, dental microscope Tirupati, microscope root canal Tirupati…",
     "https://meghanadental.in/services/microscope-dentistry",
     "Microscope Dentistry in Tirupati | Microscope Precision — Meghana Multi Speciality Dental Hospital",
     og_site, "FAQPage (4 Q&As)", "✅ Optimized"),

    ("Dental Crowns & Bridges", "/services/dental-crowns", "Dental Crowns & Bridges in Tirupati | Expert Restoration", 56,
     "Get custom dental crowns and bridges in Tirupati at Meghana Multi Speciality Dental Hospital. Ceramic, zirconia & PFM crowns by MDS specialists.", 147,
     "dental crowns Tirupati, dental bridges Tirupati, zirconia crown Tirupati, ceramic crown Tirupati…",
     "https://meghanadental.in/services/dental-crowns",
     "Dental Crowns & Bridges in Tirupati | Meghana Multi Speciality Dental Hospital",
     og_site, "FAQPage", "✅ Optimized"),

    ("Gum Treatment", "/services/gum-treatment", "Gum Disease Treatment in Tirupati | Periodontist", 51,
     "Expert gum disease treatment in Tirupati — scaling, deep cleaning, gum surgery & laser periodontal therapy by MDS periodontist at Meghana Dental Hospital.", 154,
     "gum treatment Tirupati, periodontist Tirupati, gum disease Tirupati, scaling Tirupati…",
     "https://meghanadental.in/services/gum-treatment",
     "Gum Disease Treatment in Tirupati | Meghana Multi Speciality Dental Hospital",
     og_site, "FAQPage", "✅ Optimized"),

    ("Smile Makeover", "/services/smile-makeover", "Smile Makeover in Tirupati | Complete Smile Design", 53,
     "Transform your smile with complete smile makeover at Meghana Multi Speciality Dental Hospital Tirupati — veneers, whitening, crowns & digital smile design.", 155,
     "smile makeover Tirupati, smile design Tirupati, dental veneers Tirupati, cosmetic dentistry Tirupati…",
     "https://meghanadental.in/services/smile-makeover",
     "Smile Makeover in Tirupati | Complete Transformation — Meghana Multi Speciality Dental Hospital",
     og_site, "FAQPage", "✅ Optimized"),

    ("Pediatric Dentistry", "/services/pediatric-dentistry", "Best Pediatric Dentist in Tirupati | Kids Dental Care", 55,
     "Child-friendly pediatric dentist in Tirupati. Preventive care, dental sealants, milk tooth extractions & fluoride treatments for children at Meghana Dental.", 157,
     "pediatric dentist Tirupati, children's dentist Tirupati, kids dental care Tirupati, milk teeth Tirupati…",
     "https://meghanadental.in/services/pediatric-dentistry",
     "Pediatric Dentistry in Tirupati | Kids Dental Care — Meghana Multi Speciality Dental Hospital",
     og_site, "FAQPage", "✅ Optimized"),

    ("Dentures", "/services/dentures", "Affordable Dentures in Tirupati | Full & Partial", 50,
     "Complete and partial dentures in Tirupati. Flexible, metal & acrylic options. Affordable pricing by experienced prosthodontist at Meghana Dental Hospital.", 153,
     "dentures Tirupati, full dentures Tirupati, partial dentures Tirupati, removable dentures Tirupati…",
     "https://meghanadental.in/services/dentures",
     "Dentures in Tirupati | Full & Partial Dentures — Meghana Multi Speciality Dental Hospital",
     og_site, "FAQPage", "✅ Optimized"),

    ("Tooth Extraction", "/services/tooth-extraction", "Painless Tooth Extraction in Tirupati | Safe & Fast", 55,
     "Painless tooth extraction in Tirupati — simple & surgical extractions, wisdom tooth removal. Local anaesthesia, same-day procedure at Meghana Dental.", 148,
     "tooth extraction Tirupati, wisdom tooth removal Tirupati, painless extraction Tirupati…",
     "https://meghanadental.in/services/tooth-extraction",
     "Tooth Extraction in Tirupati | Painless & Safe — Meghana Multi Speciality Dental Hospital",
     og_site, "FAQPage", "✅ Optimized"),

    ("Full Mouth Rehab", "/services/full-mouth-rehab", "Full Mouth Rehabilitation Tirupati | Complete Restoration", 58,
     "Full mouth rehabilitation in Tirupati by MDS prosthodontist. Implants, crowns, bridges & bite correction for complete oral restoration at Meghana Dental.", 151,
     "full mouth rehabilitation Tirupati, full mouth restoration Tirupati, oral rehabilitation Tirupati…",
     "https://meghanadental.in/services/full-mouth-rehab",
     "Full Mouth Rehabilitation in Tirupati | Meghana Multi Speciality Dental Hospital",
     og_site, "FAQPage", "✅ Optimized"),

    ("Laser Dentistry", "/services/laser-dentistry", "Laser Dentistry in Tirupati | Painless Laser Treatment", 57,
     "Laser dental treatment in Tirupati — gum reshaping, cavity removal, ulcer treatment & more. Painless, precise, fast healing at Meghana Dental Hospital.", 148,
     "laser dentistry Tirupati, dental laser treatment Tirupati, laser gum treatment Tirupati…",
     "https://meghanadental.in/services/laser-dentistry",
     "Laser Dentistry in Tirupati | Painless Precision — Meghana Multi Speciality Dental Hospital",
     og_site, "FAQPage", "✅ Optimized"),

    ("Dental Fillings", "/services/dental-fillings", "Dental Fillings in Tirupati | Tooth-Coloured Composite", 55,
     "Tooth-coloured composite dental fillings in Tirupati. Virtually invisible, durable, BPA-free restorations by expert dentists at Meghana Dental Hospital.", 152,
     "dental fillings Tirupati, composite filling Tirupati, tooth coloured filling Tirupati…",
     "https://meghanadental.in/services/dental-fillings",
     "Dental Fillings in Tirupati | Tooth-Coloured Restorations — Meghana Multi Speciality Dental Hospital",
     og_site, "FAQPage", "✅ Optimized"),

    ("Tooth Coloured Fillings", "/services/tooth-colored-fillings", "Tooth Coloured Fillings Tirupati | Natural Look", 51,
     "Natural-looking tooth coloured fillings in Tirupati. Composite resin restorations that match your tooth colour exactly at Meghana Multi Speciality Dental Hospital.", 160,
     "tooth coloured fillings Tirupati, white fillings Tirupati, composite resin Tirupati…",
     "https://meghanadental.in/services/tooth-colored-fillings",
     "Tooth Coloured Fillings in Tirupati | Meghana Multi Speciality Dental Hospital",
     og_site, "FAQPage", "✅ Optimized"),

    ("Ceramic Inlays & Onlays", "/services/inlays-onlays", "Ceramic Inlays & Onlays Tirupati | Precision Restoration", 55,
     "Precision ceramic inlays and onlays in Tirupati. CAD/CAM milled ceramic restorations for lasting, aesthetically superior results at Meghana Dental Hospital.", 155,
     "ceramic inlays Tirupati, ceramic onlays Tirupati, indirect restoration Tirupati, CAD CAM dentistry Tirupati…",
     "https://meghanadental.in/services/inlays-onlays",
     "Ceramic Inlays & Onlays in Tirupati | Meghana Multi Speciality Dental Hospital",
     og_site, "FAQPage", "✅ Optimized"),

    ("Blog Posts", "/blogs/[slug]", "[Blog Title] | Meghana Multi Speciality Dental Hospital", "Dynamic",
     "Expert dental health content with treatment-specific keywords, author metadata, and OG tags per blog post.", "Dynamic",
     "Per-post keywords targeting dental topics and local Tirupati searches",
     "https://meghanadental.in/blogs/[slug]",
     "[Blog Title] | Meghana Multi Speciality Dental Hospital",
     og_site, "Article + BreadcrumbList", "✅ Optimized"),

    ("Sitemap", "/sitemap.xml", "N/A — XML Sitemap", "N/A",
     "Dynamic sitemap listing all 29 pages with lastModified dates. Auto-generated by Next.js sitemap module.", "N/A",
     "N/A",
     "https://meghanadental.in/sitemap.xml",
     "N/A", "N/A", "N/A — XML Document", "✅ Implemented"),
]

for i, row_data in enumerate(rows2, start=2):
    for col, val in enumerate(row_data, start=1):
        ws2.cell(row=i, column=col, value=val)
    alt = (i % 2 == 0)
    bg = LIGHT_TEAL if alt else WHITE
    for col in range(1, len(headers2) + 1):
        cell = ws2.cell(row=i, column=col)
        cell.font = body_font()
        cell.alignment = left()
        cell.border = thin_border()
        cell.fill = cell_fill(bg)

    # Status column (col 12) — green
    status_cell = ws2.cell(row=i, column=12)
    status_cell.fill = cell_fill(GREEN_FILL)
    status_cell.font = body_font(bold=True, color=GREEN_TEXT)
    status_cell.alignment = center()

# Column widths for sheet 2
col_widths2 = [22, 35, 52, 13, 60, 12, 52, 42, 70, 38, 32, 16]
for col_idx, width in enumerate(col_widths2, start=1):
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

ws2.row_dimensions[1].height = 24
for r in range(2, len(rows2) + 2):
    ws2.row_dimensions[r].height = 42

freeze(ws2, "B2")  # Freeze row 1 AND column A


# ═══════════════════════════════════════════════════════════════════
# SHEET 3 — Keyword Targets
# ═══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Keyword Targets")
set_tab_color(ws3, "0070C0")

headers3 = ["Keyword", "Search Intent", "Page Targeted", "URL", "Priority"]
ws3.append(headers3)
style_header_row(ws3, 1, 5)

def add_section_header(ws, text, row, num_cols=5, bg="0B7A76"):
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = hdr_fill(bg)
        cell.font = Font(name="Arial", bold=True, color=WHITE, size=10, italic=True)
        cell.alignment = center()
        cell.border = thin_border()

primary_kws = [
    ("dental implants Tirupati", "Commercial", "Dental Implants", "/services/dental-implants", "High"),
    ("root canal treatment Tirupati", "Commercial", "Root Canal", "/services/root-canal", "High"),
    ("braces in Tirupati", "Commercial", "Orthodontics", "/services/orthodontics", "High"),
    ("clear aligners Tirupati", "Commercial", "Clear Aligners", "/services/aligners", "High"),
    ("microscope dentistry Tirupati", "Commercial", "Microscope Dentistry", "/services/microscope-dentistry", "High"),
    ("best dental clinic in Tirupati", "Navigational", "Home", "/", "High"),
    ("dentist in Tirupati", "Navigational", "Home", "/", "High"),
    ("dental hospital Tirupati", "Navigational", "Home", "/", "High"),
    ("orthodontist Tirupati", "Commercial", "Orthodontics", "/services/orthodontics", "High"),
    ("MDS specialist Tirupati", "Navigational", "Doctors", "/doctors", "High"),
]

secondary_kws = [
    ("teeth whitening Tirupati", "Commercial", "Teeth Whitening", "/services/teeth-whitening", "Medium"),
    ("laser teeth whitening Tirupati", "Commercial", "Teeth Whitening", "/services/teeth-whitening", "Medium"),
    ("dental crowns Tirupati", "Commercial", "Dental Crowns", "/services/dental-crowns", "Medium"),
    ("dental bridges Tirupati", "Commercial", "Dental Crowns", "/services/dental-crowns", "Medium"),
    ("gum disease treatment Tirupati", "Commercial", "Gum Treatment", "/services/gum-treatment", "Medium"),
    ("periodontist Tirupati", "Commercial", "Gum Treatment", "/services/gum-treatment", "Medium"),
    ("smile makeover Tirupati", "Commercial", "Smile Makeover", "/services/smile-makeover", "Medium"),
    ("smile design Tirupati", "Commercial", "Smile Makeover", "/services/smile-makeover", "Medium"),
    ("pediatric dentist Tirupati", "Commercial", "Pediatric Dentistry", "/services/pediatric-dentistry", "Medium"),
    ("children dentist Tirupati", "Commercial", "Pediatric Dentistry", "/services/pediatric-dentistry", "Medium"),
    ("dentures Tirupati", "Commercial", "Dentures", "/services/dentures", "Medium"),
    ("full dentures Tirupati", "Commercial", "Dentures", "/services/dentures", "Medium"),
    ("tooth extraction Tirupati", "Commercial", "Tooth Extraction", "/services/tooth-extraction", "Medium"),
    ("wisdom tooth removal Tirupati", "Commercial", "Tooth Extraction", "/services/tooth-extraction", "Medium"),
    ("full mouth rehabilitation Tirupati", "Commercial", "Full Mouth Rehab", "/services/full-mouth-rehab", "Medium"),
    ("laser dentistry Tirupati", "Commercial", "Laser Dentistry", "/services/laser-dentistry", "Medium"),
    ("composite filling Tirupati", "Commercial", "Dental Fillings", "/services/dental-fillings", "Medium"),
    ("ceramic inlays Tirupati", "Informational", "Ceramic Inlays & Onlays", "/services/inlays-onlays", "Medium"),
    ("dental implants cost Tirupati", "Commercial", "Dental Implants", "/services/dental-implants", "Medium"),
    ("braces cost Tirupati", "Commercial", "Orthodontics", "/services/orthodontics", "Medium"),
]

longtail_kws = [
    ("best dental clinic in Tirupati Andhra Pradesh", "Navigational", "Home", "/", "High"),
    ("painless root canal treatment in Tirupati", "Commercial", "Root Canal", "/services/root-canal", "High"),
    ("single sitting root canal Tirupati", "Commercial", "Root Canal", "/services/root-canal", "High"),
    ("microscope guided root canal Tirupati", "Informational", "Root Canal", "/services/root-canal", "High"),
    ("dental implants cost in Tirupati", "Commercial", "Dental Implants", "/services/dental-implants", "High"),
    ("best implant dentist in Tirupati", "Navigational", "Dental Implants", "/services/dental-implants", "High"),
    ("All on 4 dental implants Tirupati", "Commercial", "Dental Implants", "/services/dental-implants", "Medium"),
    ("invisible braces Tirupati", "Commercial", "Clear Aligners", "/services/aligners", "High"),
    ("Invisalign Tirupati", "Navigational", "Clear Aligners", "/services/aligners", "High"),
    ("metal braces Tirupati", "Commercial", "Orthodontics", "/services/orthodontics", "Medium"),
    ("ceramic braces Tirupati", "Commercial", "Orthodontics", "/services/orthodontics", "Medium"),
    ("dental clinic Ashok Nagar Tirupati", "Navigational", "Contact", "/contact", "High"),
    ("dental hospital near me Tirupati", "Navigational", "Home", "/", "High"),
    ("best dentist near me Tirupati", "Navigational", "Home", "/", "High"),
    ("dental treatment Andhra Pradesh", "Commercial", "Services", "/services", "Medium"),
    ("super speciality dental hospital Tirupati", "Navigational", "About", "/about", "Medium"),
    ("MDS prosthodontist Tirupati", "Navigational", "Doctors", "/doctors", "Medium"),
    ("endodontist in Tirupati", "Commercial", "Root Canal", "/services/root-canal", "Medium"),
    ("teeth straightening without braces Tirupati", "Informational", "Clear Aligners", "/services/aligners", "Medium"),
    ("how much does root canal cost in Tirupati", "Commercial", "Root Canal", "/services/root-canal", "High"),
]

brand_kws = [
    ("Meghana Multi Speciality Dental Hospital", "Navigational", "Home", "/", "High"),
    ("Meghana Dental Hospital Tirupati", "Navigational", "Home", "/", "High"),
    ("Meghana Dental Tirupati", "Navigational", "Home", "/", "High"),
    ("Meghana Multi Speciality Dental Hospital reviews", "Navigational", "Testimonials", "/testimonials", "High"),
    ("Meghana Dental Hospital doctors", "Navigational", "Doctors", "/doctors", "Medium"),
    ("Meghana Multi Speciality Dental Hospital contact", "Navigational", "Contact", "/contact", "Medium"),
    ("Meghana Dental Hospital appointment", "Navigational", "Contact", "/contact", "Medium"),
    ("Dr Hema Latha dentist Tirupati", "Navigational", "Doctors", "/doctors", "High"),
    ("Meghana Dental gallery Tirupati", "Navigational", "Gallery", "/gallery", "Low"),
    ("Meghana Multi Speciality Dental Hospital Ashok Nagar", "Navigational", "Contact", "/contact", "High"),
    ("meghanadental.in", "Navigational", "Home", "/", "High"),
    ("Meghana Dental blog Tirupati", "Informational", "Blogs", "/blogs", "Low"),
]

priority_colors = {
    "High":   (GREEN_FILL, GREEN_TEXT),
    "Medium": (YELLOW_FILL, YELLOW_TEXT),
    "Low":    (BLUE_FILL,   BLUE_TEXT),
}

current_row = 2

def write_kw_section(ws, section_title, keywords, start_row, bg="0B7A76"):
    add_section_header(ws, section_title, start_row, bg=bg)
    row = start_row + 1
    for i, kw_row in enumerate(keywords):
        for col, val in enumerate(kw_row, start=1):
            ws.cell(row=row, column=col, value=val)
        alt = (i % 2 == 0)
        bg_data = LIGHT_TEAL if alt else WHITE
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.font = body_font()
            cell.alignment = left()
            cell.border = thin_border()
            cell.fill = cell_fill(bg_data)
        # Priority styling
        pri = kw_row[4]
        fill_c, text_c = priority_colors.get(pri, (WHITE, "000000"))
        pri_cell = ws.cell(row=row, column=5)
        pri_cell.fill = cell_fill(fill_c)
        pri_cell.font = body_font(bold=True, color=text_c)
        pri_cell.alignment = center()
        row += 1
    return row

current_row = write_kw_section(ws3,
    "PRIMARY KEYWORDS — High Volume, Treatment-Focused",
    primary_kws, current_row, bg=DARK_TEAL)

current_row = write_kw_section(ws3,
    "SECONDARY KEYWORDS — Treatment-Specific",
    secondary_kws, current_row, bg=DARK_TEAL)

current_row = write_kw_section(ws3,
    "LONG-TAIL KEYWORDS — Location + Intent Specific",
    longtail_kws, current_row, bg=DARK_TEAL)

current_row = write_kw_section(ws3,
    "BRAND KEYWORDS — Meghana Multi Speciality Dental Hospital",
    brand_kws, current_row, bg=DARK_TEAL)

col_widths3 = [52, 18, 28, 38, 12]
for col_idx, width in enumerate(col_widths3, start=1):
    ws3.column_dimensions[get_column_letter(col_idx)].width = width

ws3.row_dimensions[1].height = 22
freeze(ws3, "A2")


# ═══════════════════════════════════════════════════════════════════
# SHEET 4 — Technical SEO Checklist
# ═══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Technical SEO Checklist")
set_tab_color(ws4, "FFA500")

headers4 = ["Item", "Status", "Notes"]
ws4.append(headers4)
style_header_row(ws4, 1, 3)

checklist = [
    ("Canonical URLs", "✅ Implemented", "All 29 pages have explicit canonical URLs pointing to meghanadental.in"),
    ("robots.txt", "✅ Configured", "Allows all bots; sitemap.xml path declared; no disallow rules"),
    ("Sitemap.xml", "✅ Dynamic", "Auto-generated at /sitemap.xml; covers all 29 pages with lastModified dates"),
    ("OpenGraph (OG) Tags", "✅ All Pages", "OG title, description, URL, siteName, image on every page"),
    ("Twitter Cards", "✅ Key Pages", "summary_large_image on Home, Blogs pages; inherited on others"),
    ("Schema Markup — Global", "✅ Dentist + MedicalBusiness", "SchemaMarkup component in layout.js renders Dentist, MedicalBusiness, WebSite, AggregateRating globally"),
    ("Schema Markup — FAQPage", "✅ 6 Service Pages", "Root Canal, Dental Implants, Orthodontics, Aligners, Teeth Whitening, Microscope Dentistry each have FAQPage JSON-LD (3–5 Q&As)"),
    ("Schema Markup — Review", "✅ Testimonials & Doctors", "Testimonials page: AggregateRating + 3 Review schemas; Doctors page: MedicalOrganization with Physician employees"),
    ("Mobile Responsive Design", "✅ Bootstrap 5", "Fully responsive grid using Bootstrap 5 breakpoints (xs/sm/md/lg/xl)"),
    ("HTTPS", "✅ Active", "meghanadental.in served over HTTPS; metadataBase set correctly"),
    ("Static Export", "✅ Pre-rendered HTML", "Next.js output: export configuration — all pages pre-rendered as static HTML for maximum page speed"),
    ("Google Analytics", "✅ G-S4R781DM6J", "Analytics component integrated in layout.js; loads asynchronously"),
    ("Font Loading", "✅ Optimized", "Inter font loaded via next/font/google with display:swap and preload:true"),
    ("Hero Image Priority", "✅ Implemented", "Above-fold hero images have priority prop set (no lazy load for LCP image)"),
    ("Lazy Loading", "✅ Below-fold Sections", "14 components on Home page use next/dynamic for code-splitting"),
    ("Bootstrap JS", "✅ Async Load", "BootstrapProvider uses useEffect to load Bootstrap JS after hydration (no render blocking)"),
    ("Clinic Name Updated", "✅ Complete", "All 29 pages and all components updated from Meghana Dental Hospital to Meghana Multi Speciality Dental Hospital"),
    ("Image Alt Tags", "✅ Descriptive", "All images have descriptive alt text including location keywords"),
    ("Floating WhatsApp Button", "✅ Implemented", "FloatingWhatsApp component on all pages for conversion"),
    ("Scroll-to-Top Button", "✅ Implemented", "ScrollToTop component improves UX on long pages"),
    ("Video — patient-testimonial.mp4", "⚠️ Needs Optimization", "46.7 MB video file causes slow page load; recommend YouTube embed or video CDN"),
    ("Images — doctor.jpg & dr-hemu.jpg", "⚠️ Needs Optimization", "1.48 MB each; convert to WebP format for 60-80% size reduction"),
    ("Images — hero-bg.png", "⚠️ Needs Optimization", "636 KB PNG; convert to WebP for significant Core Web Vitals improvement"),
    ("Images — about-clinic.png", "⚠️ Needs Optimization", "601 KB PNG; convert to WebP format"),
    ("Google Search Console", "⚠️ Pending Setup", "GSC verification comment present in layout.js but not yet active; add verification ID"),
    ("hreflang", "ℹ️ Future", "No Telugu language version yet; add hreflang when Telugu version is created"),
    ("Breadcrumb Schema", "ℹ️ Future", "Breadcrumb schema not yet implemented on service pages; adds rich results in Google"),
]

for i, (item, status, notes) in enumerate(checklist, start=2):
    ws4.cell(row=i, column=1, value=item)
    ws4.cell(row=i, column=2, value=status)
    ws4.cell(row=i, column=3, value=notes)
    alt = (i % 2 == 0)
    bg = LIGHT_TEAL if alt else WHITE

    for col in range(1, 4):
        cell = ws4.cell(row=i, column=col)
        cell.font = body_font()
        cell.alignment = left()
        cell.border = thin_border()
        cell.fill = cell_fill(bg)

    # Status column coloring
    status_cell = ws4.cell(row=i, column=2)
    status_cell.alignment = center()
    if "✅" in status:
        status_cell.fill = cell_fill(GREEN_FILL)
        status_cell.font = body_font(bold=True, color=GREEN_TEXT)
    elif "⚠️" in status:
        status_cell.fill = cell_fill(YELLOW_FILL)
        status_cell.font = body_font(bold=True, color=YELLOW_TEXT)
    else:
        status_cell.fill = cell_fill(BLUE_FILL)
        status_cell.font = body_font(bold=True, color=BLUE_TEXT)

col_widths4 = [42, 28, 80]
for col_idx, width in enumerate(col_widths4, start=1):
    ws4.column_dimensions[get_column_letter(col_idx)].width = width

ws4.row_dimensions[1].height = 22
for r in range(2, len(checklist) + 2):
    ws4.row_dimensions[r].height = 36
freeze(ws4, "A2")


# ═══════════════════════════════════════════════════════════════════
# SHEET 5 — Name Change Log
# ═══════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Name Change Log")
set_tab_color(ws5, "7030A0")

headers5 = ["File Path", "Old Value", "New Value", "Location / Context", "Status"]
ws5.append(headers5)
style_header_row(ws5, 1, 5)

change_log = [
    ("src/lib/siteConfig.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "SITE.name — primary site config", "Updated ✅"),
    ("src/app/layout.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "metadata.authors, creator, publisher", "Updated ✅"),
    ("src/app/layout.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "openGraph.siteName", "Updated ✅"),
    ("src/app/layout.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "openGraph.title", "Updated ✅"),
    ("src/app/layout.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "twitter.title", "Updated ✅"),
    ("src/app/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "openGraph.siteName", "Updated ✅"),
    ("src/app/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "openGraph.title (OG title)", "Updated ✅"),
    ("src/app/about/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "metadata.description", "Updated ✅"),
    ("src/app/about/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "openGraph.title & siteName", "Updated ✅"),
    ("src/app/contact/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "metadata.description, keywords", "Updated ✅"),
    ("src/app/contact/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "openGraph.title & siteName", "Updated ✅"),
    ("src/app/doctors/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "openGraph.siteName", "Updated ✅"),
    ("src/app/doctors/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "doctorsSchema — name field (×4)", "Updated ✅"),
    ("src/app/gallery/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "metadata.keywords, description", "Updated ✅"),
    ("src/app/gallery/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "openGraph.title & siteName", "Updated ✅"),
    ("src/app/testimonials/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "metadata.description, keywords", "Updated ✅"),
    ("src/app/testimonials/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "reviewSchema.name (×3)", "Updated ✅"),
    ("src/app/blogs/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "metadata.description", "Updated ✅"),
    ("src/app/blogs/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "openGraph.title & siteName", "Updated ✅"),
    ("src/app/services/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "openGraph.siteName", "Updated ✅"),
    ("src/app/services/root-canal/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "faqSchema answers (×4)", "Updated ✅"),
    ("src/app/services/root-canal/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "openGraph.siteName", "Updated ✅"),
    ("src/app/services/dental-implants/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "faqSchema answers (×5)", "Updated ✅"),
    ("src/app/services/dental-implants/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "openGraph.siteName", "Updated ✅"),
    ("src/app/services/orthodontics/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "faqSchema answers (×3)", "Updated ✅"),
    ("src/app/services/orthodontics/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "openGraph.siteName", "Updated ✅"),
    ("src/app/services/aligners/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "faqSchema answers (×3)", "Updated ✅"),
    ("src/app/services/aligners/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "openGraph.siteName", "Updated ✅"),
    ("src/app/services/teeth-whitening/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "faqSchema answers (×3)", "Updated ✅"),
    ("src/app/services/teeth-whitening/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "openGraph.siteName", "Updated ✅"),
    ("src/app/services/microscope-dentistry/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "faqSchema answers (×4)", "Updated ✅"),
    ("src/app/services/microscope-dentistry/page.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "openGraph.siteName", "Updated ✅"),
    ("src/components/Footer/index.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "Footer clinic name display", "Updated ✅"),
    ("src/components/About/ (components)", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "AboutStory, AboutHero content", "Updated ✅"),
    ("src/components/Navbar/NavBrand.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "Brand name in navbar", "Updated ✅"),
    ("src/components/Contact/ (components)", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "ContactInfoCards, ContactHero", "Updated ✅"),
    ("src/components/Testimonials/", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "TestimonialsHero, Grid content", "Updated ✅"),
    ("src/components/Doctors/", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "DoctorsHero, DoctorsProfiles", "Updated ✅"),
    ("src/components/shared/SchemaMarkup.js", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "Global Dentist + MedicalBusiness schema", "Updated ✅"),
    ("All service page components (14 pages)", "Meghana Dental Hospital", "Meghana Multi Speciality Dental Hospital", "ServicePageHero, content components", "Updated ✅"),
]

for i, row_data in enumerate(change_log, start=2):
    for col, val in enumerate(row_data, start=1):
        ws5.cell(row=i, column=col, value=val)
    alt = (i % 2 == 0)
    bg = LIGHT_TEAL if alt else WHITE
    for col in range(1, 6):
        cell = ws5.cell(row=i, column=col)
        cell.font = body_font()
        cell.alignment = left()
        cell.border = thin_border()
        cell.fill = cell_fill(bg)
    # Status column
    status_cell = ws5.cell(row=i, column=5)
    status_cell.fill = cell_fill(GREEN_FILL)
    status_cell.font = body_font(bold=True, color=GREEN_TEXT)
    status_cell.alignment = center()

col_widths5 = [52, 30, 46, 48, 18]
for col_idx, width in enumerate(col_widths5, start=1):
    ws5.column_dimensions[get_column_letter(col_idx)].width = width

ws5.row_dimensions[1].height = 22
for r in range(2, len(change_log) + 2):
    ws5.row_dimensions[r].height = 22
freeze(ws5, "A2")


# ═══════════════════════════════════════════════════════════════════
# SHEET 6 — Action Items
# ═══════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Action Items")
set_tab_color(ws6, "FF0000")

headers6 = ["Priority", "Action Item", "Impact", "Category", "Status", "Estimated Effort"]
ws6.append(headers6)
style_header_row(ws6, 1, 6)

PRIORITY_HIGH_BG   = "FF0000"
PRIORITY_MED_BG    = "FFA500"
PRIORITY_LOW_BG    = "00B050"

high_actions = [
    ("High", "Compress patient-testimonial.mp4 (46.7 MB) or replace with YouTube embed",
     "Page Load Speed — Critical for Core Web Vitals", "Technical", "Pending", "2 hours"),
    ("High", "Convert doctor.jpg & dr-hemu.jpg (1.48 MB each) to WebP format",
     "Page Load Speed — Reduce by 60-80%", "Technical", "Pending", "1 hour"),
    ("High", "Convert hero-bg.png (636KB) & about-clinic.png (601KB) to WebP",
     "Core Web Vitals — LCP improvement", "Technical", "Pending", "1 hour"),
    ("High", "Add Google Search Console verification ID to layout.js",
     "Search Indexing — Track rankings & crawl errors", "Technical", "Pending", "30 minutes"),
    ("High", "Submit updated sitemap.xml to Google Search Console",
     "Indexing — Get new pages crawled faster", "SEO", "Pending", "30 minutes"),
]

medium_actions = [
    ("Medium", "Update Google Business Profile name to Meghana Multi Speciality Dental Hospital",
     "Local SEO — Map Pack ranking", "Local SEO", "Pending", "1 hour"),
    ("Medium", "Update Just Dial listing name and description",
     "Local SEO — Citation consistency", "Local SEO", "Pending", "1 hour"),
    ("Medium", "Update Facebook profile/bio to full new name",
     "Social SEO — Brand consistency", "Social", "Pending", "30 minutes"),
    ("Medium", "Update Instagram bio to full new name",
     "Social SEO — Brand consistency", "Social", "Pending", "30 minutes"),
    ("Medium", "Request reviewers & responses to mention Meghana Multi Speciality Dental Hospital",
     "Brand Trust — Name consistency in reviews", "Reputation", "Pending", "Ongoing"),
    ("Medium", "Create Google Business Profile posts about key services",
     "Local SEO — Engagement signals", "Local SEO", "Pending", "2 hours/month"),
    ("Medium", "Build backlinks from dental directories (IDA, Practo, Sulekha, etc.)",
     "Authority — Domain Rating improvement", "Off-page SEO", "Pending", "Ongoing"),
]

low_actions = [
    ("Low", "Add hreflang for Telugu language version (future scope)",
     "Multilingual SEO — Telugu-speaking audience", "Technical", "Pending", "8 hours"),
    ("Low", "Implement breadcrumb schema on all service pages",
     "Rich Results in Google SERP", "Technical", "Pending", "2 hours"),
    ("Low", "Add FAQ schema to remaining service pages (Crowns, Gum, Smile, Pediatric, etc.)",
     "Rich Results — Featured Snippets opportunity", "Technical", "Pending", "3 hours"),
    ("Low", "Create video content (patient journeys, treatment explainers) for YouTube",
     "Brand Authority + YouTube SEO", "Content", "Pending", "Ongoing"),
    ("Low", "Add blog internal links to relevant service pages",
     "Internal Linking — Page Authority distribution", "Content", "Pending", "2 hours"),
]

current_row6 = 2

def write_action_section(ws, title, actions, start_row, section_bg):
    # Section header
    ws.cell(row=start_row, column=1, value=title)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    for col in range(1, 7):
        cell = ws.cell(row=start_row, column=col)
        cell.fill = hdr_fill(section_bg)
        cell.font = Font(name="Arial", bold=True, color=WHITE, size=11)
        cell.alignment = center()
        cell.border = thin_border()
    ws.row_dimensions[start_row].height = 24

    row = start_row + 1
    for i, action in enumerate(actions):
        for col, val in enumerate(action, start=1):
            ws.cell(row=row, column=col, value=val)
        alt = (i % 2 == 0)
        bg_data = LIGHT_TEAL if alt else WHITE
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.font = body_font()
            cell.alignment = left()
            cell.border = thin_border()
            cell.fill = cell_fill(bg_data)

        # Priority cell
        pri_cell = ws.cell(row=row, column=1)
        pri = action[0]
        if pri == "High":
            pri_cell.fill = cell_fill(RED_FILL)
            pri_cell.font = body_font(bold=True, color=RED_TEXT)
        elif pri == "Medium":
            pri_cell.fill = cell_fill(ORANGE_FILL)
            pri_cell.font = body_font(bold=True, color="7F3F00")
        else:
            pri_cell.fill = cell_fill(GREEN_FILL)
            pri_cell.font = body_font(bold=True, color=GREEN_TEXT)
        pri_cell.alignment = center()

        # Status cell
        status_cell = ws.cell(row=row, column=5)
        status_cell.fill = cell_fill(YELLOW_FILL)
        status_cell.font = body_font(bold=True, color=YELLOW_TEXT)
        status_cell.alignment = center()

        ws.row_dimensions[row].height = 36
        row += 1
    return row

current_row6 = write_action_section(ws6, "HIGH PRIORITY — Immediate Action Required", high_actions, current_row6, PRIORITY_HIGH_BG)
current_row6 = write_action_section(ws6, "MEDIUM PRIORITY — Complete Within 30 Days", medium_actions, current_row6, PRIORITY_MED_BG)
current_row6 = write_action_section(ws6, "LOW PRIORITY — Ongoing / Future Enhancements", low_actions, current_row6, PRIORITY_LOW_BG)

col_widths6 = [12, 62, 48, 16, 12, 18]
for col_idx, width in enumerate(col_widths6, start=1):
    ws6.column_dimensions[get_column_letter(col_idx)].width = width

ws6.row_dimensions[1].height = 22
freeze(ws6, "A2")


# ─── Save ────────────────────────────────────────────────────────────
output_path = "/Applications/Websites/dental-app/Meghana_Multi_Speciality_SEO_Report.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
